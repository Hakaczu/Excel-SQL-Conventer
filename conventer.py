from openpyxl import load_workbook
 
# config
table_name = 'Table Name'
create_stmt = False
first_row_as_col_names = True
dest_file_name = 'Excel.xlsx'
output_file_name = 'create.sql'
#end config

def create_table(table_name, columns):
    create = 'CREATE TABLE ' + table_name + '(id INT AUTO_INCREMENT PRIMARY KEY, '
    col_len = len(columns)
    col = 1
    for column in columns:
        if(col == col_len):
            stmt = column + ' VARCHAR(2000)'
        else:
            stmt = column + ' VARCHAR(2000), '
        create += stmt
        col =+ 1
    create += ');\n'
    return create

def get_columns_names(sheet, first_row_as_col_names):
    columns = []

    if first_row_as_col_names == True:
        for x in range(1, sheet.max_column + 1):
            value = str(sheet.cell(row = 1, column = x).value)
            columns.append(value)
    else:
        for x in range(1, sheet.max_column + 1):
            columns.append(str('Col' + x))
    return columns

def insert(table_name, columns, values):
    insert = 'INSERT INTO ' + table_name + '('
    col_len = len(columns)
    col = 1
    for column in columns:
        if(col == col_len):
            insert += column
        else:
            insert += column + ', '
        col += 1
    insert += ') VALUES('
    col = 1
    for val in values:
        if(col==col_len):
            insert += "'"+ val +"'"
        else: 
            insert += "'"+ val +"',"
        col += 1
    insert += ');\n'
    return insert

if __name__ == "__main__":
    workbook = load_workbook(filename = dest_file_name)
    worksheet = workbook.active
    max_col = worksheet.max_column
    max_row = worksheet.max_row
    convert = ''

    print("MAX COL: " + str(max_col))
    print("MAX ROW: " + str(max_row))

    columns_names = get_columns_names(worksheet, first_row_as_col_names)

    if create_stmt == True:
        create = create_table(table_name, columns_names)
        convert += create

    if first_row_as_col_names == True:
        
        for row in range(2, max_row + 1):
            val = []
            for col in range (1, max_col + 1):
                print(str(worksheet.cell(row = row, column = col).value))
                val.append(str(worksheet.cell(row = row, column = col).value))
            ins = insert(table_name, columns_names, val)
            convert += ins

    f = open(output_file_name, "w")
    f.write(convert)
    f.close()